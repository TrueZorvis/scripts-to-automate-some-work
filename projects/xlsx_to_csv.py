#! Python 3
# Экспортирует данные сотрудника компании в csv-файл
# (для импорта в адресную книгу почтового клиента Thunderbird)

import openpyxl
import csv

wb = openpyxl.load_workbook('Телефонный справочник.xlsx')  # читаем excel-файл
wb.active = 0  # делаем первый лист активным
sheet = wb.active  # получаем активный лист

rows = sheet.max_row
cols = sheet.max_column

need_departments = {'Отдел #1',
                    'Департамент #1',
                    'Бюро #1',
                    'Отдел #2',
                    'Цех #1',
                    'Управление #1',
                    'Отдел #3',
                    'Отдел #4', }
company = 'CompanyName'
all_employees = []
continue_export = False

for i in range(5, rows + 1):

    if sheet.cell(row=i, column=1).value in need_departments and sheet.cell(row=i, column=6).value is None:
        continue_export = True
        cell_department = sheet.cell(row=i, column=1)
        continue
    elif sheet.cell(row=i, column=1).value not in need_departments and sheet.cell(row=i, column=6).value is None:
        continue_export = False

    if continue_export:
        mylist = []
        cell_position = sheet.cell(row=i, column=1)
        cell_fullname = sheet.cell(row=i, column=2)
        lastname = str(cell_fullname.value).split(' ')[0]
        firstname = str(cell_fullname.value).split(' ')[1]
        fathersname = str(cell_fullname.value).split(' ')[2]
        cell_email = sheet.cell(row=i, column=6)

        mylist.append(cell_fullname.value)  # ФИО
        mylist.append(firstname)  # Имя
        mylist.append(fathersname)  # Отчество
        mylist.append(lastname)  # Фамилия
        mylist.append(company)  # Организация
        mylist.append(cell_position.value)  # Должность
        mylist.append(cell_department.value)  # Категория
        mylist.append(cell_email.value)  # Электронная почта
        all_employees.append(mylist)  # Добавить строку с данными работника в общий лист работников

with open('thunderbird_adressbook.csv', 'w', newline='', encoding='utf-8') as csv_file:
    writer = csv.writer(csv_file, delimiter=',')
    for employee in all_employees:
        writer.writerow(employee)

print('Экспорт адресной книги завершен')
